# =====================================================
# TEMPPLAZZA v3
# PART 1
# LOGIN + ADMIN + BASIC SETTINGS
# =====================================================

import streamlit as st
from pathlib import Path
import json
from datetime import datetime



# =====================================================
# APP INFORMATION
# =====================================================


COMPANY_NAME = "AJB - TAB"

CREATOR_NAME = "Sudhin@2026"



# =====================================================
# SESSION STATE INITIALIZATION
# =====================================================


if "username" not in st.session_state:

    st.session_state.username = ""


if "report_user" not in st.session_state:

    st.session_state.report_user = ""


if "logged_in" not in st.session_state:

    st.session_state.logged_in = False



# =====================================================
# REPORT USER CONTROL
# =====================================================


login_user = st.session_state.get(

    "username",

    ""

).lower()



# =====================================================
# ADMIN CAN CREATE REPORT FOR OTHER USERS
# =====================================================


if login_user == "admin":


    report_user = st.text_input(

        "Daily Report User Name",

        key="report_user_input"

    )


    if report_user:


        st.session_state.report_user = report_user



# =====================================================
# NORMAL USER REPORT
# =====================================================


elif login_user:


    st.session_state.report_user = (

        st.session_state.username

    )


else:


    st.session_state.report_user = ""



# =====================================================
# PAGE SETTINGS
# =====================================================


st.set_page_config(

    page_title="AJB - TAB",

    page_icon="🌡️",

    layout="wide"

)



# =====================================================
# FOLDER SETTINGS
# =====================================================


BASE_FOLDER = Path(__file__).parent


STATE_FILE = BASE_FOLDER / "TempPlazza_state.json"



# =====================================================
# LOGIN STATE
# =====================================================


def load_state():


    if STATE_FILE.exists():


        with open(
            STATE_FILE,
            "r"
        ) as file:


            return json.load(file)



    return {


        "logged_in": False,


        "username": ""


    }




def save_state():


    data = {


        "logged_in":
        st.session_state.logged_in,


        "username":
        st.session_state.username


    }


    with open(
        STATE_FILE,
        "w"
    ) as file:


        json.dump(
            data,
            file,
            indent=4
        )





# =====================================================
# SESSION START
# =====================================================


if "loaded" not in st.session_state:


    old = load_state()


    st.session_state.logged_in = old["logged_in"]


    st.session_state.username = old["username"]


    st.session_state.loaded = True





# =====================================================
# USER DATABASE
# =====================================================


ADMIN_USER = "admin"


if "users" not in st.session_state:


    st.session_state.users = {


        "admin":
        "admin@123"


    }





# =====================================================
# SIDEBAR
# =====================================================


with st.sidebar:



    st.title(
        "☰ TempPlazza Menu"
    )



    st.write(

        f"User : {st.session_state.username}"

    )



    st.divider()



    # -------------------------
    # EXCEL PLACE HOLDER
    # -------------------------


    st.subheader(
        "Excel"
    )


    st.info(
        "Excel will appear after update"
    )




    # -------------------------
    # REPORT PLACE HOLDER
    # -------------------------


    st.subheader(
        "Reports"
    )



    if st.button(
        "📄 Generate PDF",
        key="side_pdf"
    ):


        st.info(
            "PDF module coming in next part"
        )





    # -------------------------
    # ADMIN PANEL
    # -------------------------


    if (

        st.session_state.logged_in

        and

        st.session_state.username == ADMIN_USER

    ):



        st.divider()



        st.subheader(
            "🔐 Admin Panel"
        )



        # ADD USER


        new_user = st.text_input(

            "New Username",

            key="admin_add_user"

        )



        new_password = st.text_input(

            "New Password",

            type="password",

            key="admin_add_password"

        )




        if st.button(

            "➕ Add User",

            key="admin_add_button"

        ):



            if new_user and new_password:



                if new_user in st.session_state.users:


                    st.error(
                        "User already exists"
                    )


                else:


                    st.session_state.users[new_user] = new_password


                    st.success(
                        "User added"
                    )


            else:


                st.error(
                    "Enter username and password"
                )






        # REMOVE USER


        remove_user = st.text_input(

            "Remove Username",

            key="admin_remove_user"

        )




        if st.button(

            "❌ Remove User",

            key="admin_remove_button"

        ):



            if (

                remove_user in st.session_state.users

                and

                remove_user != ADMIN_USER

            ):


                del st.session_state.users[remove_user]


                st.success(
                    "User removed"
                )


            else:


                st.error(
                    "Cannot remove admin or invalid user"
                )







        # CHANGE PASSWORD



        change_user = st.text_input(

            "Change Password User",

            key="admin_change_user"

        )



        change_password = st.text_input(

            "New Password",

            type="password",

            key="admin_change_password"

        )





        if st.button(

            "🔑 Update Password",

            key="admin_password"

        ):



            if change_user in st.session_state.users:



                st.session_state.users[change_user] = change_password


                st.success(
                    "Password updated"
                )


            else:


                st.error(
                    "User not found"
                )






    st.divider()



    # LOGOUT



    if st.button(

        "🚪 Logout",

        key="final_logout"

    ):



        st.session_state.logged_in = False


        st.session_state.username = ""


        save_state()


        st.rerun()






    # -------------------------
    # FOOTER
    # -------------------------



    st.markdown(
    """
    <br><br>

    **AJB - TAB**

    Created by:  
    Sudhin@2026

    """,
    unsafe_allow_html=True
    )






# =====================================================
# LOGIN PAGE
# =====================================================



if not st.session_state.logged_in:



    st.title(
        COMPANY_NAME
    )



    st.subheader(
        "Temperature Data Logger"
    )



    username = st.text_input(
        "Username"
    )



    password = st.text_input(

        "Password",

        type="password"

    )





    if st.button(

        "🔐 Login",

        key="login"

    ):



        if (

            username in st.session_state.users

            and

            st.session_state.users[username] == password

        ):



            st.session_state.logged_in = True


            st.session_state.username = username


            save_state()


            st.rerun()



        else:


            st.error(
                "Invalid username or password"
            )



    st.write(

        f"Created by {CREATOR_NAME}"

    )



    st.stop()





# =====================================================
# AFTER LOGIN
# DASHBOARD HOME SCREEN
# =====================================================


st.title(
    "🌡️ TempPlazza Dashboard"
)


st.success(
    f"Welcome {st.session_state.username}"
)



st.divider()



# =====================================================
# DASHBOARD COUNTERS
# =====================================================


col1, col2, col3, col4 = st.columns(4)



with col1:

    st.metric(
        "Total Equipment",
        "0"
    )



with col2:

    st.metric(
        "Completed Today",
        "0"
    )



with col3:

    st.metric(
        "Pending Test",
        "0"
    )



with col4:

    st.metric(
        "Reports Created",
        "0"
    )



st.divider()



# =====================================================
# PROJECT INFORMATION
# =====================================================


st.subheader(
    "🏢 Project Information"
)



c1, c2, c3 = st.columns(3)



with c1:

    st.text_input(
        "Project Name",
        placeholder="Enter Project Name",
        key="dashboard_project"
    )



with c2:

    st.text_input(
        "Tower",
        placeholder="Enter Tower",
        key="dashboard_tower"
    )



with c3:

    st.text_input(
        "Level",
        placeholder="Enter Level",
        key="dashboard_level"
    )



st.divider()



# =====================================================
# QUICK ACTIONS
# =====================================================


st.subheader(
    "⚡ Quick Actions"
)



q1, q2, q3, q4 = st.columns(4)



with q1:

    if st.button(
        "📋 Equipment Entry",
        use_container_width=True
    ):

        st.info(
            "Equipment Entry Module - Coming in Part 2"
        )



with q2:

    if st.button(
        "📊 Daily Report",
        use_container_width=True
    ):

        st.info(
            "Daily Report Module - Coming later"
        )



with q3:

    if st.button(
        "📄 PDF Report",
        use_container_width=True
    ):

        st.info(
            "PDF Module - Coming later"
        )



with q4:

    if st.button(
        "⬇ Excel",
        use_container_width=True
    ):

        st.info(
            "Excel Module - Coming in Part 2"
        )



st.divider()



# =====================================================
# RECENT ACTIVITY
# =====================================================


st.subheader(
    "📝 Recent Activity"
)



activity_data = [

    {
        "Date":
        datetime.now().date(),

        "User":
        st.session_state.username,

        "Activity":
        "Login"

    }

]



st.table(
    activity_data
)



st.divider()



# =====================================================
# FOOTER
# =====================================================


st.markdown(
"""
<br>

<center>

<b>AJB - TAB</b><br>

Created by: Sudhin@2026

</center>

""",
unsafe_allow_html=True
)
# =====================================================
# PART 2A
# TEMPPLAZZA WORKBOOK ENGINE
# =====================================================


from openpyxl import load_workbook
from openpyxl.cell.cell import MergedCell
import shutil



# =====================================================
# FILE SETTINGS
# =====================================================


MASTER_FILE = BASE_FOLDER / "TempPlazza.xlsx"


MAIN_OUTPUT = BASE_FOLDER / "Main_Report"


MAIN_OUTPUT.mkdir(
    exist_ok=True
)



# =====================================================
# SESSION VARIABLES
# =====================================================


if "level_rows" not in st.session_state:

    st.session_state.level_rows = {}



if "last_saved_key" not in st.session_state:

    st.session_state.last_saved_key = ""



if "tower" not in st.session_state:

    st.session_state.tower = ""



if "level" not in st.session_state:

    st.session_state.level = ""



if "system" not in st.session_state:

    st.session_state.system = ""





# =====================================================
# MERGED CELL SAFE WRITE
# =====================================================


def write_merged(ws, address, value):


    cell = ws[address]


    if isinstance(cell, MergedCell):


        for merged in ws.merged_cells.ranges:


            if address in merged:


                real = ws.cell(

                    row=merged.min_row,

                    column=merged.min_col

                )


                real.value = value

                return



    else:

        cell.value = value






# =====================================================
# HIDE TEMPLATE
# =====================================================


def hide_template_sheet(wb):


    for ws in wb.worksheets:


        if ws.title.lower() in [

            "sheet1",

            "template",

            "master"

        ]:


            ws.sheet_state = "hidden"



    return wb






# =====================================================
# LOCK / UNLOCK LEVEL
# =====================================================


def lock_level(ws):


    ws.protection.sheet = True

    ws.protection.password = "TempPlazza"




def unlock_level(ws):


    ws.protection.sheet = False







# =====================================================
# CREATE TOWER FILE
# =====================================================


def get_tower_file():


    tower_file = (

        MAIN_OUTPUT /

        f"{st.session_state.tower}.xlsx"

    )



    if not tower_file.exists():


        shutil.copy(

            MASTER_FILE,

            tower_file

        )


    return tower_file







# =====================================================
# OPEN TOWER
# =====================================================


def open_tower_excel():


    return load_workbook(

        get_tower_file()

    )







# =====================================================
# CREATE LEVEL SHEET
# =====================================================


def get_level_sheet():


    wb = open_tower_excel()



    level_name = st.session_state.level




    if level_name not in wb.sheetnames:



        template = wb.worksheets[0]



        new_sheet = wb.copy_worksheet(

            template

        )



        new_sheet.title = level_name





    ws = wb[level_name]



    hide_template_sheet(wb)



    return wb, ws
    # =====================================================
# PART 2B
# EQUIPMENT ENTRY + LEVEL PAGE MANAGEMENT
# =====================================================


st.subheader(
    "🏢 Project Information"
)



col1,col2,col3 = st.columns(3)



with col1:

    st.session_state.tower = st.text_input(

        "Tower / Podium"

    )



with col2:

    st.session_state.level = st.text_input(

        "Level"

    )



with col3:

    st.session_state.system = st.text_input(

        "System"

    )



st.divider()



# =====================================================
# EQUIPMENT ENTRY
# =====================================================


st.subheader(
    "🔧 Equipment Test Entry"
)



equipment_tag = st.text_input(
    "Equipment Tag"
)


ambient_db = st.text_input(
    "Ambient DB"
)


ambient_wb = st.text_input(
    "Ambient WB"
)


room = st.text_input(
    "Room"
)


set_point = st.text_input(
    "Set Point °C"
)


design_temp = st.text_input(
    "Design Temp"
)


design_rh = st.text_input(
    "Design RH"
)


reading_time = st.text_input(
    "Reading Time"
)


indoor_db = st.text_input(
    "Indoor DB"
)


indoor_wb = st.text_input(
    "Indoor WB"
)


indoor_rh = st.text_input(
    "Indoor RH %"
)


remarks = st.text_area(
    "Remarks / Issue / Delay"
)





# =====================================================
# SAVE BUTTON
# =====================================================


if st.button(

    "💾 Save Equipment",

    key="save_equipment"

):


    if not st.session_state.tower:

        st.error(
            "Enter Tower"
        )

        st.stop()



    if not st.session_state.level:

        st.error(
            "Enter Level"
        )

        st.stop()





    current_level = st.session_state.level




    # ---------------------------------
    # CREATE LEVEL ROW COUNTER
    # ---------------------------------


    if current_level not in st.session_state.level_rows:


        st.session_state.level_rows[current_level] = 17




    row = st.session_state.level_rows[current_level]




    save_key = (

        current_level +

        str(row) +

        equipment_tag

    )




    # ---------------------------------
    # DUPLICATE PROTECTION
    # ---------------------------------


    if save_key == st.session_state.last_saved_key:


        st.warning(

            "Duplicate save prevented"

        )

        st.stop()





    wb, ws = get_level_sheet()






    # ---------------------------------
    # HEADER DATA
    # ---------------------------------


    write_merged(

        ws,

        "G7",

        st.session_state.tower

    )


    write_merged(

        ws,

        "G9",

        current_level

    )


    write_merged(

        ws,

        "G11",

        st.session_state.system

    )


    write_merged(

        ws,

        "AG35",

        datetime.now().date()

    )





    # ---------------------------------
    # EQUIPMENT DATA
    # ---------------------------------


    write_merged(ws,f"A{row}",equipment_tag)


    write_merged(ws,f"J{row}",ambient_db)


    write_merged(ws,f"N{row}",ambient_wb)


    write_merged(ws,f"R{row}",room)


    write_merged(ws,f"V{row}",set_point)


    write_merged(ws,f"AB{row}",design_temp)


    write_merged(ws,f"AF{row}",design_rh)


    write_merged(ws,f"AJ{row}",reading_time)


    write_merged(ws,f"AN{row}",indoor_db)


    write_merged(ws,f"AR{row}",indoor_wb)


    write_merged(ws,f"AV{row}",indoor_rh)


    write_merged(ws,f"AZ{row}",remarks)





    # ---------------------------------
    # ONLY ONE SAVE
    # ---------------------------------


    wb.save(

        get_tower_file()

    )



    st.session_state.last_saved_key = save_key



    st.success(

        f"Saved {current_level} Row {row}"

    )



    # move next row

    st.session_state.level_rows[current_level] += 1






    # =================================================
    # CREATE NEXT PAGE FOR SAME LEVEL
    # =================================================


    if st.session_state.level_rows[current_level] > 33:



        page = 2



        while True:


            new_level = (

                f"{current_level}_Page{page}"

            )


            temp_wb = open_tower_excel()



            if new_level not in temp_wb.sheetnames:

                break



            page += 1




        st.session_state.level = new_level



        st.session_state.level_rows[new_level] = 17



        st.success(

            f"Created {new_level}"

        )





# =====================================================
# LOCK CURRENT LEVEL BUTTON
# =====================================================


if st.button(

    "🔒 Lock Current Level",

    key="lock_level_button"

):


    if st.session_state.level:


        wb,ws = get_level_sheet()


        lock_level(ws)


        wb.save(

            get_tower_file()

        )


        st.success(

            f"{st.session_state.level} Locked"

        )
# =====================================================
# PART 3A
# DAILY REPORT SYSTEM
# =====================================================


from openpyxl import Workbook, load_workbook
from pathlib import Path
import shutil

from openpyxl.styles import Font, Border, Side, Alignment
from openpyxl.utils import get_column_letter


# =====================================================
# REPORT SETTINGS
# =====================================================


REPORT_FOLDER = BASE_FOLDER / "Daily_Report"


REPORT_FOLDER.mkdir(
    exist_ok=True
)



CREATOR_NAME = "Sudhin@2026"
# =====================================================
# REMOVE OLD DAILY REPORT SESSION VARIABLE
# =====================================================

if "daily_report" in st.session_state:

    del st.session_state["daily_report"]
# =====================================================
# UNIQUE REPORT NAME
# =====================================================


def get_report_name():


    today = datetime.now().strftime(
        "%Y-%m-%d"
    )


    username = st.session_state.report_user



    base_name = (

        f"{username}_{today}"

    )



    file = REPORT_FOLDER / (

        base_name + ".xlsx"

    )



    count = 1



    while file.exists():


        file = REPORT_FOLDER / (

            f"{base_name} ({count}).xlsx"

        )


        count += 1



    return file
# =====================================================
# CREATE DAILY REPORT
# =====================================================


def create_daily_report():


    report_file = get_report_name()


    wb_new = Workbook()


    ws_new = wb_new.active


    ws_new.title = "Daily Report"
    # =====================================================
    # REPORT SHEET HEADER FORMAT
    # =====================================================


    ws_new.merge_cells(
        "A1:M1"
    )


    ws_new["A1"] = "TEMPPLAZZA DAILY REPORT"


    ws_new["A1"].font = Font(
        bold=True,
        size=16
    )


    ws_new["A1"].alignment = Alignment(
        horizontal="center"
    )



    ws_new["A2"] = (
        "Report User: "
        + st.session_state.report_user
    )


    ws_new["A3"] = (
        "Report Date: "
        + str(datetime.now().date())
    )



    # =====================================================
    # TABLE HEADER
    # =====================================================


    ws_new.append(

        [

        "No",

        "Date",

        "User",

        "Tower",

        "Level",

        "Equipment",

        "Room",

        "Set Point",

        "Reading Time",

        "Indoor DB",

        "Indoor WB",

        "Indoor RH",

        "Remarks"

        ]

    )



    # =====================================================
    # SERIAL NUMBER START
    # =====================================================


    serial_no = 1



    # =====================================================
    # READ ALL TOWER FILES
    # =====================================================


    for tower_file in MAIN_OUTPUT.glob("*.xlsx"):


        wb = load_workbook(

            tower_file,

            data_only=True

        )


        for sheet in wb.sheetnames:


            if sheet.lower() in [

                "sheet1",

                "template",

                "master"

            ]:

                continue



            ws = wb[sheet]



            for r in range(

                17,

                34

            ):


                equipment = ws[f"A{r}"].value



                if equipment:


                    ws_new.append(

                        [

                        serial_no,

                        ws["AG35"].value,

                        st.session_state.report_user,

                        ws["G7"].value,

                        sheet,

                        equipment,

                        ws[f"R{r}"].value,

                        ws[f"V{r}"].value,

                        ws[f"AJ{r}"].value,

                        ws[f"AN{r}"].value,

                        ws[f"AR{r}"].value,

                        ws[f"AV{r}"].value,

                        ws[f"AZ{r}"].value

                        ]

                    )


                    serial_no += 1



    # =====================================================
    # SAVE REPORT
    # =====================================================


    wb_new.save(

        report_file

    )


    return report_file
# =====================================================
# DAILY REPORT BUTTON
# =====================================================


st.divider()


st.subheader(
    "📊 Daily Report"
)



if st.button(

    "📄 Create Daily Report",

    key="create_daily_report_button"

):


    report_file = create_daily_report()


    st.session_state["daily_report_file"] = str(report_file)


    st.success(

        "Daily Report Created"

    )





if "daily_report_file" in st.session_state:


    report_path = Path(

        st.session_state["daily_report_file"]

    )


    if report_path.exists():


        with open(

            report_path,

            "rb"

        ) as f:


            st.download_button(

                label="⬇️ Download Editable Excel Report",

                data=f,

                file_name=report_path.name,

                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",

                key="download_daily_report"

            )